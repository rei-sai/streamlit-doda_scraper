from selenium import webdriver # 各種必要機能のダウンロード
from selenium.webdriver.chrome import service as fs
from selenium.webdriver import ChromeOptions
from webdriver_manager.core.os_manager import ChromeType
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import logging # ログ機能
import time # 待機機能
from selenium.webdriver.support.ui import WebDriverWait # 待機機能
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import datetime # 日時の自動取得機能
import pandas as pd # Excelデータ操作用
import re # Excelデータ整形用
import io  # メモリ上のファイル操作用
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import streamlit as st # streamlit用

col1, col2, col3 = st.columns(3) # 3つのカラムに分割

# ログ表示用コンテナ
log_container = st.empty()

# ログ表示関数
def display_log(message):
    log_container.write(f"現在の状況: {message}")

def load_config():
    st.text_input("メールアドレス", key="email")
    st.text_input("パスワード",type="password", key="password")
    st.text_input("検索タイトル",key="search_title")
    st.number_input("開始ページ", min_value=1, value=1, key="first_page")
    st.number_input("終了ページ", min_value=1, value=1, key="last_page")
    st.text_input("出力ファイル名", "doda_job_data",key="output_file_prefix")

def initialize_driver(): # WebDriverを初期化する

    options = Options()

    # option設定を追加（設定する理由はメモリの削減）
    options.add_argument("--headless")
    options.add_argument('--disable-gpu')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')

    # webdriver_managerによりドライバーをインストール
    # chromiumを使用したいのでchrome_type引数でchromiumを指定しておく
    CHROMEDRIVER = ChromeDriverManager(chrome_type=ChromeType.CHROMIUM).install()
    service = fs.Service(CHROMEDRIVER)
    driver = webdriver.Chrome(
                              options=options,
                              service=service
                             )
    return driver

def login_to_doda(email,password, driver): # ログイン処理
    try:
        display_log("スクレイピング処理を開始します")
        driver.get('https://doda.jp/')
        driver.find_element(By.XPATH, '//*[@id="__next"]/div[1]/div/header/div/div/div[2]/div/a[4]/button').click()

        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input#mailAddress'))).send_keys(email)
        driver.find_element(By.CSS_SELECTOR, 'input#password').send_keys(password)
        driver.find_element(By.NAME, 'doLogin').click() # メールアドレスとパスワードを入力してログイン

        WebDriverWait(driver, 5).until( # ログイン後に検索条件ページへのリンクが出ているか確認
            EC.presence_of_element_located((By.XPATH, '//*[@id="__next"]/div[2]/main/ul/li[3]/a'))
        )
        display_log("ログイン完了")

    except Exception as e:
        display_log(f"ログイン処理に失敗しました{e}")
        exit()

def navigate_to_search_conditions(password,search_title, driver): # 詳細検索を行う
    try:
        display_log("詳細検索のための条件を取得します")
        driver.find_element(By.XPATH, '//*[@id="__next"]/div[2]/main/ul/li[3]/a').click()
        driver.find_element(By.XPATH, '//*[@id="__next"]/div[2]/main/div[3]/div[2]/div[2]/a').click()
        driver.find_element(By.XPATH, f"//div[contains(@class, 'box35')]//p[contains(@class, 'mb05 bold fs16')]//a[contains(text(), '{search_title}')]").click()
        
    except Exception as e:
        display_log(f"通常の遷移処理で例外発生: {e}")

        try: # パスワード再入力画面の場合の処理
            password_input = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, 'input#password'))
            )
            password_input.send_keys(password)

            driver.find_element(By.CSS_SELECTOR, 'a#doCheckPasswordBtn').click()
            display_log("パスワード再入力成功")
            
            driver.find_element(By.XPATH, '//*[@id="__next"]/div[2]/main/div[3]/div[2]/div[2]/a').click() # 再び元の処理に合流
            driver.find_element(By.XPATH, f"//div[contains(@class, 'box35')]//p[contains(@class, 'mb05 bold fs16')]//a[contains(text(), '{search_title}')]").click()
        except Exception as e:
            display_log(f"パスワード再入力処理でも例外発生: {e}")
            exit()

def navigate_to_page(first_page, driver): # 取得開始するページへ飛ぶ
    try:
        base_url = driver.current_url # 基準となる最初のURLを取得
        target_url = f"{base_url}&page={first_page}"  # 飛びたいページのリンクを作成
        driver.get(target_url) # 指定したページ数のページに飛ぶ

        search_result = driver.current_url # 検索結果一覧のURLを取得
        return search_result

    except Exception as e:
        display_log(f"URLでのページ遷移に失敗しました: {e}. フォールバックとしてボタン方式を試みます。")
        try:
            # フォールバック: ボタンをクリック
            for _ in range(first_page - 1):
                next_button = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, '.NextButton-module_pager__buttonNext__kKOyc'))  
                )
                next_button.click()
            display_log(f"{first_page}ページ目にボタンで遷移しました")
        except Exception as fallback_e:
            display_log(f"ボタンでの遷移にも失敗しました: {fallback_e}")
            raise


def extract_company_links(driver): # 検索結果一覧から会社名と詳細ページリンクを取得する
    company_names = [elem.text for elem in driver.find_elements(By.CSS_SELECTOR, 'h2')]
    company_links = [elem.get_attribute("href") for elem in driver.find_elements(By.CSS_SELECTOR, 'a.jobCard-header__link')]
    return company_names, company_links

def process_job_details(driver, company_names, company_links): # 各会社の詳細ページから必要なデータを取得する
    data = []
    display_log("会社の詳細情報の取得を行っています")

    for i, link in enumerate(company_links):
        try:
            driver.get(link)
            try: # 詳細ページのタブをクリック
                WebDriverWait(driver, 2).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, 'a.jobSearchDetail-tabArea__tab__item'))
                ).click()
            except TimeoutException:
                pass
            except Exception:
                display_log(f"詳細ページタブが見つかりませんでした")

            try: # 「連絡先」を含む要素を探して抽出
                contact_text = WebDriverWait(driver, 2).until(
                    EC.visibility_of_element_located((By.XPATH, "//div[contains(@class, 'jobSearchDetail-applicationMethod__contentWrap') and contains(string(.), '連絡先')]"))
                ).text
            except Exception:
                
                time.sleep(1) # 最初の取得が失敗した場合、1秒待機後再試行
                try:
                    WebDriverWait(driver, 2).until(
                        EC.element_to_be_clickable((By.CSS_SELECTOR, 'a.jobSearchDetail-tabArea__tab__item'))
                    ).click()

                    contact_text = WebDriverWait(driver, 2).until(
                        EC.visibility_of_element_located((By.XPATH, "//div[contains(@class, 'jobSearchDetail-applicationMethod__contentWrap') and contains(string(.), '連絡先')]"))
                    ).text
                except TimeoutException:
                    contact_text = "連絡先なし"
                    pass
                except Exception:
                    contact_text = "連絡先なし"

            try: # URLを抽出
                url = driver.find_element(By.XPATH, "//a[contains(@class, 'Text-module_text__eTjgU') and contains(@class, 'jobSearchDetail-companyOverview__link')]").text.split("\n")[0]
            except Exception:
                url = "URLなし"

            data.append(dict(会社名=company_names[i], 詳細ページリンク=link, 連絡先=contact_text, 企業公式HP=url)) # データを配列に追加

            # 出力内容の確認用
            # print(company_names[i], link, contact_text, url, sep="\n", end="\n\n")

        except Exception:
            display_log(f"リンク先ページの処理でエラーが発生しました")

    return data

def go_to_next_page(current_page, driver, search_result): # 次のページへ移動する
    try:
        display_log(f"次のページへ移動します")
        driver.get(search_result)
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, f'a[aria-label="{current_page + 1}ページ"]'))
        ).click()
        search_result = driver.current_url # 検索結果一覧ページを更新
        time.sleep(3)
        return search_result
    except Exception as e:
        display_log(f"次ページが見つからないか、処理終了: {e}")
        return None

def extract_contact_info(contact_text): # 電話番号とメールアドレスを抽出する関数
    if pd.isna(contact_text):
        return pd.NA, pd.NA
    phone_pattern = r"(\d{2,4}\s*-\s*\d{2,4}\s*-\s*\d{3,4})" # 電話番号のパターン
    email_pattern = r"([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,})" # メールアドレスのパターン

    phone = re.findall(phone_pattern, contact_text)
    email = re.findall(email_pattern, contact_text)

    return (", ".join(phone) if phone else pd.NA, ", ".join(email) if email else pd.NA)

def save_to_excel(data): # データをエクセルファイルに保存する
    display_log(f"データをExcelファイルに保存します")
    df = pd.DataFrame(data) # データフレームを作成
    # 重複を削除 (会社名が同じデータの2件目以降を削除)
    df.drop_duplicates(subset=["会社名"], keep="first", inplace=True)

    df[["電話番号", "メールアドレス"]] = df["連絡先"].apply(lambda x: pd.Series(extract_contact_info(x))) # 電話番号とメールアドレスを抽出して新しい列に追加
    current_time = datetime.datetime.now().strftime('%Y年%m月%d日_%H時%M分') # 現在時刻を整形
    output_file = f"{st.session_state.output_file_prefix}_{current_time}.xlsx"
    excel_buffer = io.BytesIO()  # メモリ上のファイルバッファ
    df.to_excel(excel_buffer, index=False)
    excel_buffer.seek(0)  # バッファの先頭に移動

    try: # Excel整形を適用
        format_excel(excel_buffer)  # メモリ上のバッファを渡す
        excel_buffer.seek(0)  # 整形後のバッファを先頭に戻す

    except Exception as e:
        display_log(f"Excelの整形中にエラーが発生しました: {e}")
    
    st.download_button(
        label="Excelファイルをダウンロード",
        data=excel_buffer,
        file_name=output_file,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    display_log(f"データをエクセルファイルに保存しました。整形作業を開始します。: {output_file}")

def format_excel(excel_buffer):
    wb = load_workbook(excel_buffer) # Excelファイルを開く
    ws = wb.active

    # 各列の幅を計算して設定
    for col in ws.columns:
        total_length = 0
        non_empty_cells = 0
        col_letter = get_column_letter(col[0].column)  # 列名（A, B, C...）

        for cell in col:
            try:
                if cell.row > 1 and cell.value:  # 見出し以外の非空セル
                    cell_length = len(str(cell.value))
                    total_length += cell_length
                    non_empty_cells += 1
            except:
                pass

        # 平均幅を計算（非空セルが存在する場合）
        if non_empty_cells > 0:
            avg_length = total_length / non_empty_cells
            adjusted_width = avg_length + 7  # 平均値に余裕を追加
        else:
            adjusted_width = 10  # デフォルト幅

        ws.column_dimensions[col_letter].width = adjusted_width

    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # 見出し行のフォーマット
    for cell in ws[1]: 
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    hyperlink_columns = ["詳細ページリンク", "企業公式HP"]  # ハイパーリンク化する列名を指定
    for col_name in hyperlink_columns:
        col_index = None
        for col in ws.iter_cols(min_row=1, max_row=1):  # 列名を探索
            if col[0].value == col_name:
                col_index = col[0].column
                break
        
        if col_index:  # 列が見つかった場合
            for cell in ws.iter_cols(min_col=col_index, max_col=col_index, min_row=2):
                for item in cell:
                    if item.value and isinstance(item.value, str) and item.value.startswith("http"):  # URLっぽい値
                        item.hyperlink = item.value  # ハイパーリンク設定
                        item.style = "Hyperlink"  # スタイル適用

    for row_index, row in enumerate(ws.iter_rows(min_row=2), start=2): # データ行の色付け（交互にグレー）
        if row_index % 2 == 0:
            fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            for cell in row:
                cell.fill = fill

    output_buffer = io.BytesIO()  # 整形後のExcelファイルを保存するためのバッファ
    wb.save(output_buffer)
    output_buffer.seek(0)
    excel_buffer.seek(0)
    excel_buffer.write(output_buffer.read()) # バッファの内容を上書き
    excel_buffer.truncate() # 元のバッファの長さを超えないようにする
    wb.close()


def main(config): # 上記の処理をまとめて行うメイン関数

    driver = initialize_driver()

    try:
        login_to_doda(st.session_state.email, st.session_state.password, driver)
        navigate_to_search_conditions(st.session_state.password,st.session_state.search_title, driver)
        search_result = navigate_to_page(st.session_state.first_page, driver)
        data = []
        current_page = config["first_page"]
        last_page = config["last_page"]

        while current_page <= last_page:

            company_names, company_links = extract_company_links(driver)
            data.extend(process_job_details(driver, company_names, company_links))
            new_search_result = go_to_next_page(current_page, driver, search_result)

            if not new_search_result:
                break
            search_result = new_search_result
            current_page += 1

        save_to_excel(data)

    finally:
        driver.quit()
        display_log(f"WebDriverを終了しました{e}")

# st.session_state の初期化
if "start_config" not in st.session_state:
    st.session_state.start_config = False
if "start_scraping" not in st.session_state:
    st.session_state.start_scraping = False

with col2: # 中央のカラムにボタンを配置
# 「スクレイピング設定開始」ボタン
    if st.button("検索条件の設定を開始する"):
        st.session_state.start_config = True
        st.session_state.start_scraping = False

# 設定入力フォーム
if st.session_state.start_config and not st.session_state.start_scraping:
    load_config()
    # 「スクレイピング実行」ボタン
    if st.button("上記の設定で処理を実行"):
        st.session_state.start_scraping = True
        # 設定値をconfigに格納
        st.session_state.config = {
            "email": st.session_state.email,
            "password": st.session_state.password,
            "search_title": st.session_state.search_title,
            "first_page": st.session_state.first_page,
            "last_page": st.session_state.last_page,
            "output_file_prefix": st.session_state.output_file_prefix
        }

if st.session_state.start_scraping and st.session_state.config:
    main(st.session_state.config) # st.session_state.configをmain関数に渡す
    st.session_state.start_scraping = False
    st.session_state.start_config = False
